"""CSV/Excel export builders, profit-margin calc, and download-link generator."""
import base64
import io
from datetime import datetime

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font

from cogs.formatters import format_cost


def calculate_profit_margins(cost_per_box, sales_price_per_box):
    if sales_price_per_box <= 0:
        return {"profit_per_box": 0, "profit_margin_percent": 0, "roi_percent": 0}

    profit_per_box = sales_price_per_box - cost_per_box
    profit_margin_percent = (profit_per_box / sales_price_per_box) * 100 if sales_price_per_box > 0 else 0
    roi_percent = (profit_per_box / cost_per_box) * 100 if cost_per_box > 0 else 0

    return {
        "profit_per_box": profit_per_box,
        "profit_margin_percent": profit_margin_percent,
        "roi_percent": roi_percent,
    }


def create_csv_export(summary_data_dict, profit_data, calculation_details):
    display_currency = st.session_state.get("display_currency", "USD")
    export_data = []

    export_data.append(["Calculation Details"])
    export_data.append(["Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    export_data.append(["Product", calculation_details.get('product', 'N/A')])
    export_data.append(["Quantity", calculation_details.get('quantity', 'N/A')])
    export_data.append(["Shipment Type", calculation_details.get('shipment_type', 'N/A')])
    export_data.append(["Fixed Cost Mode", calculation_details.get('fixed_cost_mode', 'N/A')])
    export_data.append([])

    export_data.append(["Cost Breakdown"])
    for key, value in summary_data_dict.items():
        if value is not None:
            export_data.append([key, value])
    export_data.append([])

    if profit_data:
        export_data.append(["Profit Analysis"])
        export_data.append(["Cost per Box (USD)", f"${profit_data.get('final_cost_per_box_usd', 0):,.2f}"])
        export_data.append([f"Cost per Box ({display_currency})", format_cost(profit_data.get('final_cost_per_box_usd', 0))])

    return pd.DataFrame(export_data)


def create_excel_export(summary_data_dict, profit_data, calculation_details, sensitivity_data):
    display_currency = st.session_state.get("display_currency", "USD")
    wb = Workbook()

    if wb.active is not None:
        wb.remove(wb.active)

    ws_summary = wb.create_sheet("Cost Summary")
    ws_summary.append(["Cost Calculator - Summary Report"])
    ws_summary.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws_summary.append([])

    ws_summary.append(["Calculation Details"])
    ws_summary.append(["Product", calculation_details.get('product', 'N/A')])
    ws_summary.append(["Quantity", calculation_details.get('quantity', 'N/A')])
    ws_summary.append(["Shipment Type", calculation_details.get('shipment_type', 'N/A')])
    ws_summary.append([])

    ws_summary.append(["Cost Breakdown"])
    for key, value in summary_data_dict.items():
        if value is not None:
            ws_summary.append([key, value])

    if profit_data:
        ws_profit = wb.create_sheet("Profit Analysis")
        ws_profit.append(["Profit Analysis"])
        ws_profit.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws_profit.append([])
        ws_profit.append(["Cost per Box (USD)", f"${profit_data.get('final_cost_per_box_usd', 0):,.2f}"])
        ws_profit.append([f"Cost per Box ({display_currency})", format_cost(profit_data.get('final_cost_per_box_usd', 0))])

        if sensitivity_data:
            ws_profit.append([])
            ws_profit.append(["Sensitivity Analysis"])
            headers = list(sensitivity_data[0].keys())
            ws_profit.append(headers)
            for row in sensitivity_data:
                ws_profit.append([row[header] for header in headers])

    return wb


def season_matrix_to_excel(long_df, *, price_basis="Cost", built_at="", inputs_df=None) -> bytes:
    """Render the Season Pricing long table to .xlsx bytes for st.download_button.

    Sheet "Season pricing" is the matrix with a bold, frozen header row and a
    currency format on the per-pallet columns. Sheet "Run info" records the price
    basis, build timestamp, and (optionally) the raw price behind each fruit so
    the export is self-documenting. UI-free so it is unit-testable.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Season pricing"

    cols = list(long_df.columns)
    ws.append(cols)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for _, row in long_df.iterrows():
        ws.append([row[c] for c in cols])

    price_idx = [i for i, c in enumerate(cols, start=1) if str(c).endswith("pallets")]
    for ci in price_idx:
        for ri in range(2, ws.max_row + 1):
            ws.cell(row=ri, column=ci).number_format = '"$"#,##0.00'

    ws.freeze_panes = "A2"
    for ci, c in enumerate(cols, start=1):
        body_widths = [len(str(v)) for v in long_df[c]] if len(long_df) else []
        width = max([len(str(c))] + body_widths)
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = min(max(width + 2, 10), 42)

    info = wb.create_sheet("Run info")
    info.append(["Season pricing export"])
    info["A1"].font = Font(bold=True)
    info.append(["Price basis", price_basis])
    info.append(["Built", built_at])
    info.append(["Rows", len(long_df)])
    if inputs_df is not None and len(inputs_df):
        info.append([])
        info.append(list(inputs_df.columns))
        for cell in info[info.max_row]:
            cell.font = Font(bold=True)
        for _, row in inputs_df.iterrows():
            info.append([row[c] for c in inputs_df.columns])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def get_download_link(data, filename, file_type):
    if file_type == "csv":
        csv = data.to_csv(index=False, header=False)
        b64 = base64.b64encode(csv.encode()).decode()
        return f'<a href="data:file/csv;base64,{b64}" download="{filename}">Download CSV</a>'
    elif file_type == "excel":
        buffer = io.BytesIO()
        data.save(buffer)
        buffer.seek(0)
        b64 = base64.b64encode(buffer.read()).decode()
        return f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">Download Excel</a>'
    return ""
