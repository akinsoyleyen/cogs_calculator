# -*- coding: utf-8 -*-
import pandas as pd
import streamlit as st
import requests
import os
from datetime import datetime
import plotly.express as px
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import base64
import numpy as np  # Add this import at the top for np.unique
import plotly.graph_objects as go

# --- Constants ---
INTEREST_RATE = 0.05  # 5% interest rate
INTEREST_COST_ITEM_NAME = "Calc. Interest"

# Everything in the app is stored and computed in USD. A single display FX is
# applied at the end to convert results into any chosen reporting currency.
DISPLAY_CURRENCIES = ["USD", "EUR", "GBP", "TRY", "CAD", "CHF", "JPY", "AED", "SGD", "AUD"]
CURRENCY_SYMBOLS = {
    "USD": "$", "EUR": "€", "GBP": "£", "TRY": "₺", "CAD": "C$",
    "CHF": "CHF ", "JPY": "¥", "AED": "د.إ ", "SGD": "S$", "AUD": "A$"
}
FALLBACK_USD_RATES = {  # USD -> target, used when Frankfurter fails
    "USD": 1.0, "EUR": 0.90, "GBP": 0.78, "TRY": 38.00, "CAD": 1.38,
    "CHF": 0.90, "JPY": 148.0, "AED": 3.67, "SGD": 1.34, "AUD": 1.52
}

# Frankfurter supports a single base→target lookup
FRANKFURTER_URL_TEMPLATE = "https://api.frankfurter.app/latest?from=USD&to={target}"

# Legacy aliases (some code still reads these — kept for backward compatibility)
exchange_rate = 1.0  # TRY rate no longer used; retained as a no-op multiplier

# --- File Paths ---
COMPONENTS_CSV = "components.csv"
RECIPE_CSV = "product_recipe.csv"
FIXED_CSV = "fixed_costs.csv"
WEIGHTS_CSV = "product_weights.csv"
AIR_RATES_CSV = "air_freight_rates.csv"
PALLETS_CSV = "pallets.csv"
PACKING_CSV = "product_packing.csv"

# --- Streamlit Setup ---
st.set_page_config(layout="wide", page_title="Ledger — Cost & Logistics", page_icon="◐")

# --- Inject styling (fonts, palette, typography) ---
CUSTOM_CSS = """
<style>
@import url('https://fonts.googleapis.com/css2?family=Space+Grotesk:wght@400;500;600;700&family=JetBrains+Mono:wght@400;500;600&display=swap');
:root {
  /* Ink on paper — minimalist, inspired by SwingScope Light */
  --paper:   #fafaf7;
  --paper-2: #f4f3ee;
  --paper-3: #ebe9e1;
  --card:    #ffffff;
  --rule:      rgba(20, 16, 30, 0.08);
  --rule-soft: rgba(20, 16, 30, 0.14);
  --ink:       #15121f;
  --ink-soft:  #3a3545;
  --ink-muted: #6c6478;
  --ink-faint: #a8a0b0;
  /* Restrained accents — pink is the single brand signal; green/red for up/down data only */
  --pink:   #e11d74;
  --amber:  #b87d00;
  --violet: #6b3fd4;
  --cyan:   #0b8f7e;
  --up:     #0b8f7e;
  --down:   #c23b3b;
  --space-1: 4px;  --space-2: 8px;  --space-3: 12px;
  --space-4: 16px; --space-5: 24px; --space-6: 32px;
  --space-7: 48px; --space-8: 64px;
}

html, body, [data-testid="stAppViewContainer"] {
  background: var(--paper) !important;
  color: var(--ink);
  font-family: "Space Grotesk", ui-sans-serif, system-ui, -apple-system, "Segoe UI", sans-serif;
  font-feature-settings: "tnum", "ss01";
  letter-spacing: -0.005em;
}

/* Hide Streamlit chrome we don't want */
#MainMenu, footer, header [data-testid="stToolbar"] { visibility: hidden; }
[data-testid="stHeader"] { background: transparent; }

/* Top masthead title */
h1, [data-testid="stAppViewContainer"] h1 {
  font-family: "Space Grotesk", ui-sans-serif, system-ui, sans-serif;
  font-weight: 600;
  font-size: clamp(2rem, 3.4vw, 2.8rem);
  line-height: 1.05;
  letter-spacing: -0.03em;
  color: var(--ink);
  margin: var(--space-3) 0 var(--space-2);
}
h1 em { font-style: normal; color: var(--pink); }

h2, [data-testid="stAppViewContainer"] h2 {
  font-family: "Space Grotesk", sans-serif;
  font-weight: 600;
  font-size: 1.35rem;
  letter-spacing: -0.015em;
  color: var(--ink);
  margin-top: var(--space-6);
  margin-bottom: var(--space-3);
}

h3 {
  font-family: "Space Grotesk", sans-serif;
  font-weight: 600;
  font-size: 0.72rem;
  letter-spacing: 0.18em;
  text-transform: uppercase;
  color: var(--ink-muted);
  margin-top: var(--space-5);
  margin-bottom: var(--space-2);
}

p, span, label, li {
  font-family: "Space Grotesk", ui-sans-serif, system-ui, sans-serif;
  color: var(--ink);
}

label, [data-testid="stWidgetLabel"] p {
  font-family: "JetBrains Mono", ui-monospace, monospace !important;
  font-size: 0.68rem !important;
  font-weight: 500 !important;
  letter-spacing: 0.14em;
  text-transform: uppercase;
  color: var(--ink-muted) !important;
}

/* Top-level horizontal rules become hairlines */
hr { border: 0; border-top: 1px solid var(--rule); margin: var(--space-6) 0; }

/* Sidebar */
[data-testid="stSidebar"] {
  background: var(--card);
  border-right: 1px solid var(--rule);
}
[data-testid="stSidebar"] [data-testid="stWidgetLabel"] p { color: var(--ink-muted) !important; }
[data-testid="stSidebar"] .stSubheader, [data-testid="stSidebar"] h3 {
  font-family: "JetBrains Mono", ui-monospace, monospace !important;
  font-size: 0.62rem !important;
  font-weight: 600 !important;
  text-transform: uppercase;
  letter-spacing: 0.18em;
  color: var(--ink) !important;
  border-bottom: 1px solid var(--rule);
  padding-bottom: var(--space-2);
  margin-top: var(--space-5);
}

/* Buttons — filled ink primary, like SwingScope Light .btn-primary */
.stButton > button, .stDownloadButton > button, [data-testid="stFormSubmitButton"] button {
  font-family: "JetBrains Mono", ui-monospace, monospace;
  font-weight: 600;
  font-size: 0.78rem;
  letter-spacing: 0.14em;
  text-transform: uppercase;
  border-radius: 2px;
  border: 1px solid var(--ink);
  background: var(--ink);
  color: var(--paper);
  padding: 11px 18px;
  transition: transform 160ms cubic-bezier(0.2,0.8,0.2,1), background 160ms ease, box-shadow 160ms ease;
}
.stButton > button:hover, .stDownloadButton > button:hover, [data-testid="stFormSubmitButton"] button:hover {
  background: var(--pink);
  border-color: var(--pink);
  color: #fff;
  transform: translateY(-1px);
  box-shadow: 0 2px 10px rgba(225,29,116,0.22);
}
.stButton > button:focus, [data-testid="stFormSubmitButton"] button:focus { box-shadow: 0 0 0 3px rgba(225,29,116,0.18); outline: none; }

/* Sidebar buttons — ghost style */
[data-testid="stSidebar"] .stButton > button {
  background: transparent;
  color: var(--ink);
  border: 1px solid var(--rule-soft);
}
[data-testid="stSidebar"] .stButton > button:hover {
  background: var(--paper-2);
  color: var(--ink);
  border-color: var(--ink);
}

/* Inputs */
input, textarea, select, [data-baseweb="input"] input, [data-baseweb="select"] > div {
  font-family: "Public Sans", sans-serif !important;
  color: var(--ink) !important;
}
[data-baseweb="input"] input, [data-baseweb="select"] > div, [data-testid="stNumberInput"] input,
[data-testid="stTextInput"] input {
  background: var(--paper) !important;
  border-radius: 2px !important;
}
[data-baseweb="base-input"] { background: var(--paper) !important; }

/* Numeric inputs: tabular */
[data-testid="stNumberInput"] input { font-family: "JetBrains Mono", ui-monospace, monospace !important; font-feature-settings: "tnum"; font-size: 0.92rem; }

/* Radio & checkboxes */
.stRadio label, .stCheckbox label { font-size: 0.9rem !important; text-transform: none !important; letter-spacing: 0 !important; color: var(--ink) !important; font-weight: 400 !important; }

/* Metrics — stat-card with left-border accent (SwingScope Light) */
[data-testid="stMetric"] {
  background: var(--card);
  border: 1px solid var(--rule);
  border-left: 2px solid var(--pink);
  border-radius: 2px;
  padding: var(--space-4) var(--space-5);
  position: relative;
  transition: border-color 160ms ease;
}
[data-testid="stMetric"]:hover { border-left-color: var(--ink); }
[data-testid="stMetricLabel"] p {
  font-family: "JetBrains Mono", ui-monospace, monospace !important;
  font-size: 0.62rem !important;
  font-weight: 500 !important;
  letter-spacing: 0.18em;
  text-transform: uppercase;
  color: var(--ink-muted) !important;
}
[data-testid="stMetricValue"] {
  font-family: "JetBrains Mono", ui-monospace, monospace !important;
  font-feature-settings: "tnum", "ss01";
  font-weight: 500 !important;
  font-size: 1.6rem !important;
  letter-spacing: -0.02em;
  color: var(--ink) !important;
  margin-top: 6px;
}
[data-testid="stMetricDelta"] { font-family: "JetBrains Mono", ui-monospace, monospace; font-size: 0.72rem; color: var(--ink-muted); }

/* Tabs — understated, active = ink underline */
.stTabs [data-baseweb="tab-list"] {
  gap: 2px;
  border-bottom: 1px solid var(--rule);
  padding-bottom: 0;
}
.stTabs [data-baseweb="tab"] {
  font-family: "JetBrains Mono", ui-monospace, monospace;
  font-weight: 500;
  font-size: 0.72rem;
  letter-spacing: 0.14em;
  text-transform: uppercase;
  color: var(--ink-muted);
  background: transparent;
  padding: 12px 16px;
  border-radius: 0;
  border-bottom: 1px solid transparent;
  margin-bottom: -1px;
  transition: color 120ms ease, border-color 120ms ease;
}
.stTabs [data-baseweb="tab"]:hover { color: var(--ink); }
.stTabs [aria-selected="true"] {
  color: var(--ink) !important;
  border-bottom: 1px solid var(--ink) !important;
  background: transparent !important;
}
.stTabs [data-baseweb="tab-highlight"], .stTabs [data-baseweb="tab-border"] { display: none; }

/* Dataframes */
[data-testid="stDataFrame"], [data-testid="stDataEditor"] {
  border: 1px solid var(--rule);
  border-radius: 2px;
  font-family: "JetBrains Mono", ui-monospace, monospace;
  background: var(--card);
}
[data-testid="stDataFrame"] [role="columnheader"], [data-testid="stDataEditor"] [role="columnheader"] {
  font-family: "JetBrains Mono", ui-monospace, monospace !important;
  font-size: 0.62rem !important;
  font-weight: 600 !important;
  letter-spacing: 0.14em;
  text-transform: uppercase;
  color: var(--ink-muted) !important;
  background: var(--paper-2) !important;
}

/* Expander */
[data-testid="stExpander"] {
  border: 1px solid var(--rule);
  border-radius: 2px;
  background: var(--card);
}
[data-testid="stExpander"] summary { font-family: "Space Grotesk", sans-serif; font-weight: 600; color: var(--ink); }

/* Alerts — flat, minimalist */
[data-testid="stAlert"] {
  border-radius: 2px;
  border: 1px solid var(--rule);
  background: var(--card);
  font-family: "Space Grotesk", sans-serif;
  color: var(--ink);
  padding: var(--space-4) var(--space-5);
}
[data-testid="stAlert"][kind="info"]    { background: var(--paper-2); border-left: 2px solid var(--violet); }
[data-testid="stAlert"][kind="warning"] { background: color-mix(in oklab, var(--amber) 6%, var(--card)); border-left: 2px solid var(--amber); }
[data-testid="stAlert"][kind="error"]   { background: color-mix(in oklab, var(--down) 6%, var(--card));  border-left: 2px solid var(--down); }
[data-testid="stAlert"][kind="success"] { background: color-mix(in oklab, var(--up) 6%, var(--card));    border-left: 2px solid var(--up); }

/* Plotly chart background */
.js-plotly-plot .plotly .bg { fill: var(--card) !important; }

/* Captions */
.stCaption, [data-testid="stCaptionContainer"], [data-testid="stCaption"] {
  font-family: "Space Grotesk", sans-serif;
  font-size: 0.78rem;
  color: var(--ink-muted);
}

/* Links — pink underline, SwingScope Light highlighter */
a { color: var(--ink); text-decoration: none; background-image: linear-gradient(transparent 62%, rgba(225,29,116,0.18) 62%); padding: 0 2px; border: none; transition: background-image 160ms ease; }
a:hover { background-image: linear-gradient(transparent 0%, rgba(225,29,116,0.22) 0%); color: var(--pink); }

/* Radio groups — stack with breathing room */
.stRadio [role="radiogroup"] { gap: var(--space-2); }

/* Columns gap rhythm */
[data-testid="column"] { padding: 0 var(--space-3); }

/* Reduce motion */
@media (prefers-reduced-motion: reduce) {
  * { transition-duration: 0.01ms !important; animation-duration: 0.01ms !important; }
}
</style>
"""
if hasattr(st, "html"):
    st.html(CUSTOM_CSS)
else:
    st.markdown(CUSTOM_CSS, unsafe_allow_html=True)

# --- Sidebar masthead ---
with st.sidebar:
    try:
        st.image("assets/Logo.png", width=96)
    except FileNotFoundError:
        st.error("Logo file not found. Make sure 'Logo.png' is in assets/.")
    except Exception as e:
        st.error(f"An error occurred loading the logo: {e}")
    st.markdown(
        "<div style='font-family:\"Space Grotesk\",sans-serif;font-weight:600;"
        "font-size:1.05rem;color:var(--ink);letter-spacing:-0.02em;margin-top:8px;'>"
        "Ledger<span style='color:var(--pink)'>.</span></div>"
        "<div style='font-family:\"JetBrains Mono\",monospace;font-size:0.62rem;color:var(--ink-muted);"
        "letter-spacing:0.2em;text-transform:uppercase;margin-bottom:18px;'>Cost · Logistics · Margin</div>",
        unsafe_allow_html=True,
    )

# --- Masthead ---
st.markdown(
    "<div style='font-family:\"JetBrains Mono\",monospace;font-size:0.66rem;color:var(--ink-muted);"
    "letter-spacing:0.24em;text-transform:uppercase;margin-bottom:-4px;'>"
    "Landed · Cost · Calculator</div>",
    unsafe_allow_html=True,
)
st.markdown(
    "<h1 style='margin-top:4px'>Product Cost &amp; <em>Logistics</em> Calculator</h1>",
    unsafe_allow_html=True,
)
st.markdown(
    "<p style='font-family:\"Space Grotesk\",sans-serif;font-size:1rem;color:var(--ink-soft);"
    "max-width:68ch;margin-top:-4px;margin-bottom:28px;line-height:1.55;'>"
    "Everything in USD. One FX at the end — pick any currency you want to report in.</p>",
    unsafe_allow_html=True,
)

# --- Exchange Rate Functions ---
@st.cache_data(ttl=3600)
def get_usd_to_target_rate(target_currency: str):
    """Fetch live USD -> target rate from Frankfurter. Returns (rate, date|None)."""
    if target_currency == "USD":
        return 1.0, "native"
    try:
        response = requests.get(FRANKFURTER_URL_TEMPLATE.format(target=target_currency), timeout=10)
        response.raise_for_status()
        data = response.json()
        rate = data.get('rates', {}).get(target_currency)
        date = data.get('date', None)
        if rate:
            return float(rate), date
    except Exception:
        pass
    return None, None

# --- Display formatting ---
# These three module-level variables are initialised further down, after the sidebar
# lets the user pick a reporting currency. All intermediate math is in USD.
display_currency = "USD"
display_fx_rate = 1.0
display_symbol = "$"

def format_cost(usd_amount):
    """Convert a USD amount to the chosen display currency and format it."""
    if usd_amount is None:
        return f"{display_symbol}0.00"
    amt = float(usd_amount) * display_fx_rate
    return f"{display_symbol}{amt:,.2f}"

# Backward-compat shims — keep the old signatures so call sites don't have to change.
def format_cost_by_mode(usd_amount, _mode_unused=None):
    return format_cost(usd_amount)

def format_cost_usd_only(usd_amount):
    return f"${0.0 if usd_amount is None else float(usd_amount):,.2f}"

def format_cost_eur_only(usd_amount):
    # Retained for any external code paths; now returns the chosen display currency.
    return format_cost(usd_amount)

def calculate_profit_margins(cost_per_box, sales_price_per_box):
    """Calculate profit margins for given cost and sales price"""
    if sales_price_per_box <= 0:
        return {"profit_per_box": 0, "profit_margin_percent": 0, "roi_percent": 0}
    
    profit_per_box = sales_price_per_box - cost_per_box
    profit_margin_percent = (profit_per_box / sales_price_per_box) * 100 if sales_price_per_box > 0 else 0
    roi_percent = (profit_per_box / cost_per_box) * 100 if cost_per_box > 0 else 0
    
    return {
        "profit_per_box": profit_per_box,
        "profit_margin_percent": profit_margin_percent,
        "roi_percent": roi_percent
    }

# --- Export Helper Functions ---
def create_csv_export(summary_data_dict, profit_data, calculation_details):
    """Create CSV export of calculation results"""
    export_data = []
    
    # Add calculation details
    export_data.append(["Calculation Details"])
    export_data.append(["Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    export_data.append(["Product", calculation_details.get('product', 'N/A')])
    export_data.append(["Quantity", calculation_details.get('quantity', 'N/A')])
    export_data.append(["Shipment Type", calculation_details.get('shipment_type', 'N/A')])
    export_data.append(["Fixed Cost Mode", calculation_details.get('fixed_cost_mode', 'N/A')])
    export_data.append([])
    
    # Add cost breakdown
    export_data.append(["Cost Breakdown"])
    for key, value in summary_data_dict.items():
        if value is not None:
            export_data.append([key, value])
    export_data.append([])
    
    # Add profit analysis if available
    if profit_data:
        export_data.append(["Profit Analysis"])
        export_data.append(["Cost per Box (USD)", f"${profit_data.get('final_cost_per_box_usd', 0):,.2f}"])
        export_data.append([f"Cost per Box ({display_currency})", format_cost(profit_data.get('final_cost_per_box_usd', 0))])

    return pd.DataFrame(export_data)

def create_excel_export(summary_data_dict, profit_data, calculation_details, sensitivity_data):
    """Create Excel export with multiple sheets and formatting"""
    wb = Workbook()
    
    # Remove default sheet
    if wb.active is not None:
        wb.remove(wb.active)
    
    # Summary sheet
    ws_summary = wb.create_sheet("Cost Summary")
    ws_summary.append(["Cost Calculator - Summary Report"])
    ws_summary.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws_summary.append([])
    
    # Add calculation details
    ws_summary.append(["Calculation Details"])
    ws_summary.append(["Product", calculation_details.get('product', 'N/A')])
    ws_summary.append(["Quantity", calculation_details.get('quantity', 'N/A')])
    ws_summary.append(["Shipment Type", calculation_details.get('shipment_type', 'N/A')])
    ws_summary.append([])
    
    # Add cost breakdown
    ws_summary.append(["Cost Breakdown"])
    for key, value in summary_data_dict.items():
        if value is not None:
            ws_summary.append([key, value])
    
    # Profit Analysis sheet
    if profit_data:
        ws_profit = wb.create_sheet("Profit Analysis")
        ws_profit.append(["Profit Analysis"])
        ws_profit.append([f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws_profit.append([])
        ws_profit.append(["Cost per Box (USD)", f"${profit_data.get('final_cost_per_box_usd', 0):,.2f}"])
        ws_profit.append([f"Cost per Box ({display_currency})", format_cost(profit_data.get('final_cost_per_box_usd', 0))])
        
        # Add sensitivity analysis
        if sensitivity_data:
            ws_profit.append([])
            ws_profit.append(["Sensitivity Analysis"])
            headers = list(sensitivity_data[0].keys())
            ws_profit.append(headers)
            for row in sensitivity_data:
                ws_profit.append([row[header] for header in headers])
    
    return wb

def get_download_link(data, filename, file_type):
    """Generate download link for files"""
    if file_type == "csv":
        csv = data.to_csv(index=False, header=False)
        b64 = base64.b64encode(csv.encode()).decode()
        href = f'<a href="data:file/csv;base64,{b64}" download="{filename}">Download CSV</a>'
    elif file_type == "excel":
        buffer = io.BytesIO()
        data.save(buffer)
        buffer.seek(0)
        b64 = base64.b64encode(buffer.read()).decode()
        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">Download Excel</a>'
    return href

# --- Initialize session state ---
if 'calculation_done' not in st.session_state: st.session_state['calculation_done'] = False
if 'final_cost_per_box_usd' not in st.session_state: st.session_state['final_cost_per_box_usd'] = 0.0


# --- Display Currency (single FX at the end) ---
st.sidebar.subheader("Reporting Currency")
display_currency = st.sidebar.selectbox(
    "Display results in:",
    DISPLAY_CURRENCIES,
    index=0,
    help="All inputs are in USD. Pick any currency here to convert the final output — one rate, applied at the end."
)
display_symbol = CURRENCY_SYMBOLS.get(display_currency, display_currency + " ")

fx_mode = st.sidebar.radio(
    "FX rate source",
    ("Live (Frankfurter API)", "Manual"),
    key="fx_mode",
    index=0,
    disabled=(display_currency == "USD")
)

if display_currency == "USD":
    display_fx_rate = 1.0
    fx_source = "native"
    fx_date = None
elif fx_mode == "Live (Frankfurter API)":
    live_rate, fx_date = get_usd_to_target_rate(display_currency)
    if live_rate is not None:
        display_fx_rate = live_rate
        fx_source = f"API ({fx_date})"
    else:
        st.sidebar.warning(f"Live rate unavailable for USD→{display_currency}. Enter manually:")
        display_fx_rate = st.sidebar.number_input(
            f"USD→{display_currency} rate",
            min_value=0.000001,
            value=float(FALLBACK_USD_RATES.get(display_currency, 1.0)),
            step=0.0001,
            format="%.6f",
            key=f"manual_fx_fallback_{display_currency}",
        )
        fx_source = "Manual fallback"
else:
    display_fx_rate = st.sidebar.number_input(
        f"USD→{display_currency} rate",
        min_value=0.000001,
        value=float(FALLBACK_USD_RATES.get(display_currency, 1.0)),
        step=0.0001,
        format="%.6f",
        key=f"manual_fx_{display_currency}",
    )
    fx_source = "Manual"

st.sidebar.metric(
    label=f"USD → {display_currency}  ({fx_source})",
    value=f"{display_fx_rate:.4f}" if display_currency != "USD" else "1.0000",
)

if st.sidebar.button("Refresh live FX", help="Clear cache and re-fetch FX rates", disabled=(display_currency == "USD")):
    st.cache_data.clear()
    st.rerun()

# Kept as aliases so legacy code paths still work
currency_display_mode = f"{display_currency} Only"
usd_to_eur_rate = display_fx_rate if display_currency == "EUR" else FALLBACK_USD_RATES["EUR"]


# --- Initialize DataFrames BEFORE loading ---
# ... (initialization remains the same) ...
components_df_try_loaded = None; product_recipe_df = None; fixed_df_usd_loaded = None
product_weights_df = None; air_rates_df = None; pallets_df = None; product_packing_df = None # Added packing_df
components_df = None; fixed_df = None; errors = []

# --- Load Data from CSV Files (with improved interest cost handling) ---
# ... (load_csv function remains largely the same, validation improved slightly) ...
def load_csv(file_path, required_cols, numeric_cols=None, decimal_char='.', string_cols=None):
    global errors
    if not os.path.exists(file_path):
        errors.append(f"File missing: '{os.path.abspath(file_path)}'")
        return None
    try:
        df = pd.read_csv(file_path, decimal=decimal_char)
        missing = [c for c in required_cols if c not in df.columns]
        if missing: raise ValueError(f"'{file_path}' missing required columns: {missing}")

        # Basic check if dataframe is empty after loading
        if df.empty and required_cols:
            st.warning(f"Warning: File '{file_path}' loaded as empty. Check content and headers.")
            # Optionally return empty dataframe with expected columns?
            # return pd.DataFrame(columns=required_cols)

        if numeric_cols:
            for col in numeric_cols:
                if col not in df.columns: continue # Skip if numeric col isn't present (might be optional)

                # Special handling for Interest Cost row in fixed_costs.csv MonthlyCost
                is_interest_cost_row = False
                if file_path.endswith(FIXED_CSV) and col == 'MonthlyCost' and 'CostItem' in df.columns:
                     # Ensure CostItem is string before comparison
                     df['CostItem'] = df['CostItem'].astype(str).fillna('')
                     is_interest_cost_row = df['CostItem'].str.strip().str.lower() == INTEREST_COST_ITEM_NAME.lower()

                # Attempt conversion to numeric, coercing errors to NaN
                original_type = df[col].dtype
                df[col] = pd.to_numeric(df[col], errors='coerce')

                # Check for NaNs *except* in the interest cost row for MonthlyCost
                has_nan = df[col].isnull()
                if isinstance(is_interest_cost_row, pd.Series): # Ensure it's a boolean Series
                     rows_to_check_for_nan = has_nan & (~is_interest_cost_row)
                else: # If not the special case, check all rows
                     rows_to_check_for_nan = has_nan

                if rows_to_check_for_nan.any():
                    # Identify specific rows with errors if possible (more helpful message)
                    error_indices = df.index[rows_to_check_for_nan].tolist()
                    raise ValueError(f"Column '{col}' in '{file_path}' contains non-numeric values or blanks at row indices (starting 0): {error_indices[:5]}...") # Show first few errors

        if string_cols:
            for col in string_cols:
                if col in df.columns:
                    df[col] = df[col].astype(str).fillna('').str.strip() # Convert, fill NA with empty string, strip whitespace

        return df
    except Exception as e:
        errors.append(f"Error processing '{file_path}': {e}")
        return None

try: # Load all data
    components_df_try_loaded = load_csv(COMPONENTS_CSV, ['ComponentName', 'CostPerUnit', 'WeightKG'], ['CostPerUnit', 'WeightKG'])
    product_recipe_df = load_csv(RECIPE_CSV, ['ProductID', 'ComponentName', 'QuantityPerProduct'], ['QuantityPerProduct'], string_cols=['ProductID', 'ComponentName']) # Ensure ComponentName is string
    fixed_df_usd_loaded = load_csv(FIXED_CSV, ['CostItem', 'MonthlyCost', 'Category'], ['MonthlyCost'], string_cols=['Category','CostItem'])
    product_weights_df = load_csv(WEIGHTS_CSV, ['ProductID', 'NetWeightKG'], ['NetWeightKG'], decimal_char=',', string_cols=['ProductID']) # Watch decimal char if issues arise
    air_rates_df = load_csv(AIR_RATES_CSV, ['Destination', 'MinWeightKG', 'PricePerKG_USD', 'AirwayBill_USD'], ['MinWeightKG', 'PricePerKG_USD', 'AirwayBill_USD'])
    pallets_df = load_csv(PALLETS_CSV, ['PalletType', 'CostUSD', 'WeightKG'], ['CostUSD', 'WeightKG'])
    product_packing_df = load_csv(PACKING_CSV, ['ProductID', 'BoxesPerPallet'], ['BoxesPerPallet'], string_cols=['ProductID']) # Load packing data

    # Explicitly convert BoxesPerPallet to numeric to avoid type issues
    if product_packing_df is not None:
        product_packing_df['BoxesPerPallet'] = pd.to_numeric(product_packing_df['BoxesPerPallet'], errors='coerce')

    if errors: raise ValueError("Errors occurred during file loading. See details above.")

    # Process data only if loading succeeded
    components_df = components_df_try_loaded.copy()
    components_df['CostPerUnit_USD'] = components_df['CostPerUnit'] * exchange_rate
    components_df['ComponentName'] = components_df['ComponentName'].astype(str).str.strip() # Ensure string type and stripped

    fixed_df = fixed_df_usd_loaded.copy()
    fixed_df = fixed_df.rename(columns={'MonthlyCost': 'MonthlyCost_USD'})
    fixed_df['Category'] = fixed_df['Category'].astype(str).str.strip().str.title()
    fixed_df['CostItem'] = fixed_df['CostItem'].astype(str).str.strip() # Ensure string type and stripped

    # Validate fixed cost categories
    valid_categories = ['Primary', 'Secondary']
    invalid_cats_df = fixed_df[~fixed_df['Category'].isin(valid_categories)]
    if not invalid_cats_df.empty:
        invalid_cats_list = invalid_cats_df['Category'].unique().tolist()
        errors.append(f"Invalid Category values in '{FIXED_CSV}': {invalid_cats_list}. Expected 'Primary' or 'Secondary'.")

    # Ensure ProductIDs are string type for consistent merging/lookup across all relevant DFs
    if product_weights_df is not None: product_weights_df['ProductID'] = product_weights_df['ProductID'].astype(str)
    else: errors.append(f"'{WEIGHTS_CSV}' failed to load or is empty.")
    if product_recipe_df is not None: product_recipe_df['ProductID'] = product_recipe_df['ProductID'].astype(str)
    else: errors.append(f"'{RECIPE_CSV}' failed to load or is empty.")
    if product_packing_df is not None: product_packing_df['ProductID'] = product_packing_df['ProductID'].astype(str)
    # No critical error if packing is missing, handled later

    if errors: raise ValueError("Errors occurred during data processing. See details above.")

except Exception as e:
    st.error(" / ".join(errors) if errors else f"An unexpected error occurred during data loading/processing: {e}")
    st.stop() # Stop execution if critical files/processing failed

# --- Get unique lists ---
# ... (logic remains the same) ...
try:
    product_ids = sorted(product_weights_df['ProductID'].unique().tolist()) # Products must have weight defined
    air_destinations = sorted(air_rates_df['Destination'].unique().tolist()) if air_rates_df is not None else []
    pallet_types = ["None"] + sorted(pallets_df['PalletType'].unique().tolist()) if pallets_df is not None and 'PalletType' in pallets_df else ["None"]
except Exception as e: st.error(f"Could not extract lists: {e}"); product_ids = []; air_destinations = []; pallet_types = ["None"]; st.stop()


# --- Streamlit User Interface ---
st.sidebar.header("Calculation Inputs")
# Product Selection
# ... (logic remains the same) ...
if not product_ids: st.sidebar.warning("No products available."); selected_product = None
else: selected_product = st.sidebar.selectbox("Select Product:", product_ids)

# --- Quantity / Pallet Linking ---
# ... (logic remains the same) ...
st.sidebar.subheader("Quantity & Pallets")
auto_calc_boxes = st.sidebar.checkbox("Calculate Box Quantity from Pallets?", value=False)

boxes_per_pallet = 0; calculated_boxes = 0; can_auto_calc = False

if auto_calc_boxes and selected_product and product_packing_df is not None:
    packing_info = product_packing_df[product_packing_df['ProductID'] == str(selected_product)]
    if not packing_info.empty:
        boxes_per_pallet = packing_info['BoxesPerPallet'].iloc[0]
        if pd.notna(boxes_per_pallet) and boxes_per_pallet > 0: can_auto_calc = True
        else: st.sidebar.warning(f"Boxes/Pallet is zero or invalid for {selected_product}.")
    else: st.sidebar.warning(f"Packing info (Boxes/Pallet) not found for {selected_product} in '{PACKING_CSV}'."); auto_calc_boxes = False

num_pallets = st.sidebar.number_input("Number of Pallets:", min_value=0, value=1, step=1)

if auto_calc_boxes and can_auto_calc:
    calculated_boxes = int(max(1, num_pallets * boxes_per_pallet))
    quantity_input = st.sidebar.number_input( "Quantity (Boxes/Units):", min_value=1, value=calculated_boxes, step=1, disabled=True, help=f"Auto: {num_pallets} pallets * {boxes_per_pallet} boxes/pallet")
    st.sidebar.caption(f"Using {int(boxes_per_pallet)} Boxes/Pallet for {selected_product}")
else:
     quantity_input = st.sidebar.number_input("Quantity (Boxes/Units):", min_value=1, value=100, step=1, disabled=False, help="Enter manually, or check box above to calculate from pallets.")

# Pallet Type Selection
# ... (logic remains the same) ...
if not pallet_types: st.sidebar.warning("No pallet types loaded."); selected_pallet_type = "None"
else: none_index = pallet_types.index("None") if "None" in pallet_types else 0; selected_pallet_type = st.sidebar.selectbox("Select Pallet Type:", pallet_types, index=none_index)

# Raw Product Cost & Other Costs
# ... (logic remains the same) ...
st.sidebar.subheader("Costs") # Group remaining costs
raw_cost_per_kg_try = st.sidebar.number_input("Raw Product Cost per KG (USD):", min_value=0.0, value=1.0, step=0.05, format="%.3f", help="Bare fruit cost per KG, in USD. One currency for everything.")

# --- Show Fixed Cost Totals in Radio Options ---
if fixed_df is not None:
    total_primary = fixed_df[fixed_df['Category'] == 'Primary']['MonthlyCost_USD'].sum()
    total_all = fixed_df[fixed_df['Category'].isin(['Primary', 'Secondary'])]['MonthlyCost_USD'].sum()
else:
    total_primary = 0.0
    total_all = 0.0
fixed_cost_options = [
    f"All Costs (Primary+Secondary) (${'{:,.2f}'.format(total_all)})",
    f"Primary Costs Only (${'{:,.2f}'.format(total_primary)})",
    "Add 10% of total value"
]
fixed_cost_selection = st.sidebar.radio("Include Fixed Costs:", fixed_cost_options, index=0)

# Map selection to logic
if "Primary Costs Only" in fixed_cost_selection:
    fixed_categories_to_include = ['Primary']
    fixed_cost_label_suffix = "(Primary Only)"
    fixed_cost_mode = "standard"
elif "10% of total value" in fixed_cost_selection:
    fixed_categories_to_include = []  # Not used in this mode
    fixed_cost_label_suffix = "(10% of Total Value)"
    fixed_cost_mode = "percent"
else:
    fixed_categories_to_include = ['Primary', 'Secondary']
    fixed_cost_label_suffix = "(All)"
    fixed_cost_mode = "standard"

unexpected_cost_try = st.sidebar.number_input("Unexpected Costs (USD for batch):", min_value=0.0, value=0.0, step=10.0, format="%.2f")
include_variable_costs = st.sidebar.checkbox("Include Variable Component Costs in COGS?", value=True)


# Logistics Inputs
# ... (logic remains the same) ...
st.sidebar.subheader("Logistics")
shipment_types = ["Select...", "Air", "Container", "Truck"]; selected_shipment_type = st.sidebar.selectbox("Select Shipment Type:", shipment_types)
selected_destination = None
if selected_shipment_type == "Air":
    if not air_destinations: st.sidebar.warning("No air destinations loaded.")
    else: selected_destination = st.sidebar.selectbox("Select Destination (Air):", ["Select..."] + air_destinations) # Add select prompt

manual_logistics_cost_usd = 0.0
manual_logistics_cost_input = 0.0  # referenced in logistics-tab display block below
if selected_shipment_type in ["Container", "Truck"]:
    manual_logistics_cost_input = st.sidebar.number_input(
        f"Fixed {selected_shipment_type} Price (USD):",
        min_value=0.0,
        value=4000.0,
        step=50.0,
        format="%.2f",
        help="Enter the flat freight price in USD. The final output is converted to your reporting currency at the end.",
    )
    manual_logistics_cost_usd = manual_logistics_cost_input

# *** NEW: Rebate Input ***
st.sidebar.subheader("Sales Adjustments")
rebate_rate_input = st.sidebar.number_input(
    "Retailer Rebate/Fee (%):",
    min_value=0.0,
    value=0.0,        # Default to 0%
    step=0.1,
    format="%.2f",    # Allow one decimal place for percentage
    help="Enter a percentage (e.g., 5 for 5%) to be added to the Total Delivered Cost."
)


# --- Calculation Logic ---
# ... (check remains the same) ...
calculation_ready = (
    selected_product is not None and
    selected_pallet_type is not None and
    fixed_cost_selection is not None and
    selected_shipment_type != "Select..." and
    (selected_shipment_type != "Air" or (selected_destination is not None and selected_destination != "Select...")) and
    quantity_input >= 1
)

if calculation_ready and st.sidebar.button("Calculate Costs"):

    # --- Initialize variables ---
    # ... (Initialize ALL calculation variables used below to 0.0 or empty dataframes/strings) ...
    total_raw_cost_usd=0.0; raw_cost_per_box_usd=0.0; total_variable_comp_cost_usd=0.0; total_per_unit_variable_comp_cost_usd=0.0; total_pallet_cost_usd=0.0; pallet_cost_per_box_usd=0.0; total_variable_costs_incl_pallets_usd=0.0; variable_costs_incl_pallets_per_box_usd=0.0; total_allocated_fixed_cost_usd=0.0; fixed_cost_per_unit_usd=0.0; total_cogs_usd=0.0; cogs_per_box_usd=0.0; cogs_per_kg_usd=0.0; total_logistics_cost_usd=0.0; logistics_per_box_usd=0.0; logistics_per_kg_gross_usd=0.0; total_unexpected_cost_usd=0.0; unexpected_cost_per_box_usd=0.0; total_delivered_cost_usd=0.0; delivered_cost_per_box_usd=0.0; delivered_cost_per_kg_net_usd=0.0; total_packaging_weight_kg=0.0; calculated_gross_weight_kg_per_box=0.0; final_shipping_gross_weight_kg=0.0; total_pallet_weight_kg=0.0; total_net_weight_kg=0.0; freight_or_fixed_logistics_cost=0.0; fixed_logistics_price=0.0; awb_cost=0.0; logistics_rate_per_kg=0.0; interest_cost_usd=0.0;
    # *** NEW: Rebate variables ***
    rebate_percentage = 0.0; rebate_amount_usd = 0.0; final_total_cost_usd = 0.0; final_cost_per_box_usd = 0.0

    product_variable_costs_detailed = pd.DataFrame() # Ensure it's defined
    calc_errors = [] # List to hold calculation errors

    try: # Main calculation try block
        selected_product_str = str(selected_product)

        # 1. Get Pallet Specs & Cost/Weight
        cost_per_pallet_usd = 0.0; weight_per_pallet_kg = 0.0
        if selected_pallet_type != "None" and num_pallets > 0 and pallets_df is not None:
            pallet_spec = pallets_df[pallets_df['PalletType'] == selected_pallet_type]
            if not pallet_spec.empty:
                cost_per_pallet_usd = pallet_spec['CostUSD'].iloc[0]
                weight_per_pallet_kg = pallet_spec['WeightKG'].iloc[0]
            else: st.warning(f"Specs missing for pallet '{selected_pallet_type}'. Cost/Weight assumed 0.")
        total_pallet_cost_usd = num_pallets * cost_per_pallet_usd
        total_pallet_weight_kg = num_pallets * weight_per_pallet_kg
        pallet_cost_per_box_usd = total_pallet_cost_usd / quantity_input if quantity_input > 0 else 0

        # 2. Get Net Weight
        product_weight_info = product_weights_df[product_weights_df['ProductID'] == selected_product_str]
        if product_weight_info.empty: raise ValueError(f"Net Weight missing for '{selected_product_str}' in '{WEIGHTS_CSV}'.")
        net_weight_kg = product_weight_info['NetWeightKG'].iloc[0]
        if not net_weight_kg > 0: raise ValueError(f"Net Weight for '{selected_product_str}' must be > 0.")

        # 3. Raw Product Cost
        raw_cost_per_kg_usd = raw_cost_per_kg_try * exchange_rate
        raw_cost_per_box_usd = raw_cost_per_kg_usd * net_weight_kg
        total_raw_cost_usd = raw_cost_per_box_usd * quantity_input

        # 4. Variable Component Costs & Packaging Weight
        total_packaging_weight_kg = 0.0 # Initialize packaging weight

        # Always calculate packaging weight from recipe for logistics, regardless of cost inclusion
        product_recipe = product_recipe_df[product_recipe_df['ProductID'] == selected_product_str]
        if not product_recipe.empty:
            recipe_components = product_recipe['ComponentName'].unique()
            available_components = components_df['ComponentName'].unique()
            missing_in_components_df = [comp for comp in recipe_components if comp not in available_components]
            if missing_in_components_df:
                 raise ValueError(f"Component(s) in recipe for '{selected_product_str}' not found in '{COMPONENTS_CSV}': {missing_in_components_df}")

            product_variable_costs_detailed = pd.merge(
                product_recipe, components_df[['ComponentName', 'CostPerUnit_USD', 'WeightKG']],
                on='ComponentName', how='left'
            )
            if product_variable_costs_detailed.isnull().values.any():
                missing = pd.Series(product_variable_costs_detailed[product_variable_costs_detailed.isnull().any(axis=1)]['ComponentName']).unique().tolist()
                raise ValueError(f"Details missing after merge for component(s): {missing}. Check '{COMPONENTS_CSV}'.")

            product_variable_costs_detailed['QuantityPerProduct'] = pd.to_numeric(product_variable_costs_detailed['QuantityPerProduct'], errors='coerce')
            if product_variable_costs_detailed['QuantityPerProduct'].isnull().any():
                raise ValueError("Non-numeric 'QuantityPerProduct' found in recipe.")

            # Calculate Weight first
            product_variable_costs_detailed['LineItemWeightKG'] = product_variable_costs_detailed['WeightKG'] * product_variable_costs_detailed['QuantityPerProduct']
            total_packaging_weight_kg = product_variable_costs_detailed['LineItemWeightKG'].sum()

            # Calculate Cost only if included
            if include_variable_costs:
                product_variable_costs_detailed['LineItemCost_USD'] = product_variable_costs_detailed['CostPerUnit_USD'] * product_variable_costs_detailed['QuantityPerProduct']
                total_per_unit_variable_comp_cost_usd = product_variable_costs_detailed['LineItemCost_USD'].sum()
                total_variable_comp_cost_usd = total_per_unit_variable_comp_cost_usd * quantity_input
            else:
                # Ensure costs are zero if not included
                product_variable_costs_detailed['LineItemCost_USD'] = 0.0
                total_per_unit_variable_comp_cost_usd = 0.0
                total_variable_comp_cost_usd = 0.0
        else:
            # No recipe found, variable costs and packaging weight are zero
            total_per_unit_variable_comp_cost_usd = 0.0
            total_variable_comp_cost_usd = 0.0
            total_packaging_weight_kg = 0.0
            st.warning(f"No recipe found for ProductID '{selected_product_str}' in '{RECIPE_CSV}'. Variable costs & packaging weight assumed 0.")

        # Total Variable (Components + Pallets)
        total_variable_costs_incl_pallets_usd = total_variable_comp_cost_usd + total_pallet_cost_usd
        variable_costs_incl_pallets_per_box_usd = total_variable_costs_incl_pallets_usd / quantity_input if quantity_input > 0 else 0

        # --- Intermediate Weights ---
        calculated_gross_weight_kg_per_box = net_weight_kg + total_packaging_weight_kg
        total_batch_gross_weight_boxes_only = calculated_gross_weight_kg_per_box * quantity_input
        final_shipping_gross_weight_kg = total_batch_gross_weight_boxes_only + total_pallet_weight_kg

        # --- 6. Logistics Costs (Freight/Fixed ONLY) ---
        # ... (logic remains the same) ...
        freight_or_fixed_logistics_cost = 0.0; logistics_rate_per_kg = 0.0; awb_cost = 0.0; fixed_logistics_price = 0.0; logistics_cost_source = "N/A"
        if selected_shipment_type == "Air":
            if air_rates_df is not None and not air_rates_df.empty:
                 applicable_rates = air_rates_df[ (air_rates_df['Destination'] == selected_destination) & (air_rates_df['MinWeightKG'] <= final_shipping_gross_weight_kg) ].sort_values('MinWeightKG', ascending=False)
                 if not applicable_rates.empty:
                     rate_row = applicable_rates.iloc[0]; logistics_rate_per_kg = rate_row['PricePerKG_USD']; awb_cost = rate_row['AirwayBill_USD']
                     freight_or_fixed_logistics_cost = (final_shipping_gross_weight_kg * logistics_rate_per_kg) + awb_cost; logistics_cost_source = f"Air: {rate_row['MinWeightKG']}+ KG Tier"
                 else: st.warning(f"No AIR rate for {selected_destination} at {final_shipping_gross_weight_kg:.2f} KG."); logistics_cost_source = "Air: No Rate Found"; freight_or_fixed_logistics_cost = 0.0
            else: st.warning(f"Air rates file '{AIR_RATES_CSV}' missing or empty."); logistics_cost_source = "Air: Rates Missing"; freight_or_fixed_logistics_cost = 0.0
        elif selected_shipment_type in ["Container", "Truck"]:
             freight_or_fixed_logistics_cost = manual_logistics_cost_usd; fixed_logistics_price = manual_logistics_cost_usd; logistics_cost_source = f"{selected_shipment_type}: Manual Input"
        total_logistics_cost_usd = freight_or_fixed_logistics_cost
        logistics_per_box_usd = total_logistics_cost_usd / quantity_input if quantity_input > 0 else 0
        logistics_per_kg_gross_usd = total_logistics_cost_usd / final_shipping_gross_weight_kg if final_shipping_gross_weight_kg > 0 else 0

        # --- 7. Unexpected Cost ---
        # ... (logic remains the same) ...
        total_unexpected_cost_usd = unexpected_cost_try * exchange_rate
        unexpected_cost_per_box_usd = total_unexpected_cost_usd / quantity_input if quantity_input > 0 else 0

        # --- 8. Fixed Costs (10% Option Calculated Here, after Logistics) ---
        interest_base_cost = 0.0
        if fixed_cost_mode == "percent":
            total_value_for_percent = total_raw_cost_usd + total_variable_costs_incl_pallets_usd + total_logistics_cost_usd + total_unexpected_cost_usd
            fixed_cost_10_percent = total_value_for_percent * 0.10
            total_allocated_fixed_cost_usd = fixed_cost_10_percent  # Only the 10% value, do not add interest here
        else:
            # Remove debug output for production
            standard_fixed_df = fixed_df[ (fixed_df['CostItem'].str.strip().str.lower() != INTEREST_COST_ITEM_NAME.lower()) & (fixed_df['Category'].isin(fixed_categories_to_include)) ]
            total_standard_fixed_cost_usd = standard_fixed_df['MonthlyCost_USD'].sum()
            total_allocated_fixed_cost_usd = total_standard_fixed_cost_usd

        # Interest should account for the entire amount we pre-finance (raw, variable, fixed, logistics, unexpected)
        interest_base_cost = (
            total_raw_cost_usd
            + total_variable_costs_incl_pallets_usd
            + total_allocated_fixed_cost_usd
            + total_logistics_cost_usd
            + total_unexpected_cost_usd
        )
        interest_cost_usd = interest_base_cost * INTEREST_RATE
        fixed_cost_per_unit_usd = total_allocated_fixed_cost_usd / quantity_input if quantity_input > 0 else 0
        total_cogs_usd = total_raw_cost_usd + total_variable_costs_incl_pallets_usd + total_allocated_fixed_cost_usd + interest_cost_usd
        cogs_per_box_usd = total_cogs_usd / quantity_input if quantity_input > 0 else 0; total_net_weight_kg = net_weight_kg * quantity_input; cogs_per_kg_usd = total_cogs_usd / total_net_weight_kg if total_net_weight_kg > 0 else 0

        # --- 11. Total Delivered Cost (Before Rebate) ---
        # ... (logic remains the same) ...
        total_delivered_cost_usd = total_cogs_usd + total_logistics_cost_usd + total_unexpected_cost_usd
        delivered_cost_per_box_usd = total_delivered_cost_usd / quantity_input if quantity_input > 0 else 0
        delivered_cost_per_kg_net_usd = total_delivered_cost_usd / total_net_weight_kg if total_net_weight_kg > 0 else 0

        # *** NEW: 12. RETAILER REBATE/FEE CALCULATION ***
        rebate_percentage = rebate_rate_input # Get percentage from sidebar input
        # Calculate the rebate amount based on the total delivered cost
        rebate_amount_usd = total_delivered_cost_usd * (rebate_percentage / 100.0)
        # Calculate the final total cost including the rebate
        final_total_cost_usd = total_delivered_cost_usd + rebate_amount_usd
        # Calculate the final cost per box including the rebate
        final_cost_per_box_usd = final_total_cost_usd / quantity_input if quantity_input > 0 else 0

        # *** NEW: 13. PROFIT MARGIN CALCULATIONS ***
        # Store profit analysis data
        st.session_state['profit_analysis'] = {
            'final_cost_per_box_usd': final_cost_per_box_usd,
            'cogs_per_box_usd': cogs_per_box_usd,
            'delivered_cost_per_box_usd': delivered_cost_per_box_usd
        }

        # --- STORE RESULTS IN SESSION STATE ---
        # Store values needed potentially by other pages (like Batch Sales, Profit Calc)
        st.session_state['last_calc_product'] = selected_product_str
        st.session_state['last_calc_quantity'] = quantity_input
        st.session_state['cogs_per_box_usd'] = cogs_per_box_usd # COGS per box
        st.session_state['unexpected_cost_per_box_usd'] = unexpected_cost_per_box_usd
        st.session_state['calculated_gross_weight_kg_per_box'] = calculated_gross_weight_kg_per_box
        st.session_state['air_rates_df'] = air_rates_df # Pass air rates table if needed elsewhere
        st.session_state['selected_shipment_type'] = selected_shipment_type
        st.session_state['delivered_cost_per_box_usd'] = delivered_cost_per_box_usd # Cost BEFORE rebate
        # *** NEW: Store final cost per box including rebate ***
        st.session_state['final_cost_per_box_usd'] = final_cost_per_box_usd # Cost AFTER rebate
        st.session_state['last_calc_rebate_rate'] = rebate_percentage # Store rebate info too
        st.session_state['last_calc_rebate_amount'] = rebate_amount_usd

        # Update summary data dictionary to include rebate and final total
        summary_data = {
            "1. Raw Product": format_cost_by_mode(total_raw_cost_usd, currency_display_mode),
            "2. Variable Costs (incl. Pallets)": format_cost_by_mode(total_variable_costs_incl_pallets_usd, currency_display_mode),
            f"3. Fixed Costs {fixed_cost_label_suffix}": (
                format_cost_by_mode(fixed_cost_10_percent, currency_display_mode)
                if fixed_cost_mode == "percent"
                else format_cost_by_mode(total_allocated_fixed_cost_usd, currency_display_mode)
            ),
            f"   (Calc. Interest @ 5.0%)": format_cost_by_mode(interest_cost_usd, currency_display_mode),
            "   Subtotal COGS": format_cost_by_mode(total_cogs_usd, currency_display_mode),
            f"4. {selected_shipment_type} Freight/Fee": format_cost_by_mode(freight_or_fixed_logistics_cost, currency_display_mode),
            "   Subtotal Logistics": format_cost_by_mode(total_logistics_cost_usd, currency_display_mode),
            "5. Unexpected Costs": format_cost_by_mode(total_unexpected_cost_usd, currency_display_mode),
            "      Delivered Cost (Before Rebate)": format_cost_by_mode(total_delivered_cost_usd, currency_display_mode),
            f"6. Rebate/Fee ({rebate_percentage:.1f}%)": format_cost_by_mode(rebate_amount_usd, currency_display_mode),
            "Grand Total Cost (incl Rebate)": format_cost_by_mode(final_total_cost_usd, currency_display_mode)
        }
        # Remove None values from summary (like zero interest) for cleaner display
        st.session_state['summary_data'] = {k: v for k, v in summary_data.items() if v is not None}

        st.session_state['calculation_done'] = True
        st.session_state['last_calc_product'] = selected_product
        st.session_state['last_calc_quantity'] = quantity_input
        st.session_state['cogs_per_box_usd'] = cogs_per_box_usd
        st.session_state['unexpected_cost_per_box_usd'] = unexpected_cost_per_box_usd
        st.session_state['calculated_gross_weight_kg_per_box'] = calculated_gross_weight_kg_per_box
        st.session_state['air_rates_df'] = air_rates_df
        st.session_state['selected_shipment_type'] = selected_shipment_type
        st.session_state['final_cost_per_box_usd'] = final_cost_per_box_usd
        st.session_state['selected_pallet_type'] = selected_pallet_type
        st.session_state['boxes_per_pallet'] = boxes_per_pallet if boxes_per_pallet > 0 else 1  # Always set, fallback to 1 if not valid
        st.session_state['weight_per_pallet_kg'] = weight_per_pallet_kg
        st.session_state['num_pallets'] = num_pallets
        # OVERWRITE THE CLEANED DICTIONARY WITH THE ORIGINAL ONE
        st.session_state['summary_data'] = summary_data

    # --- Exception Handling for Main Calculation ---
    except ValueError as ve: calc_errors.append(f"Data Input Error: {ve}")
    except KeyError as ke: calc_errors.append(f"Data Lookup Error: Missing key {ke}. Check file headers/content.")
    except ZeroDivisionError: calc_errors.append("Calculation Error: Division by zero (check Quantity).")
    except AssertionError as ae: calc_errors.append(f"Data Validation Failed: {ae}")
    except Exception as e_main_calc: calc_errors.append(f"Unexpected calculation error: {e_main_calc}")

    # --- Display Results OR Errors ---
    if calc_errors:
        st.error("Calculation failed. Please resolve errors:")
        for err in calc_errors: st.error(f"- {err}")
        st.session_state['calculation_done'] = False # Ensure flag is false on error
    elif st.session_state.get('calculation_done', False): # Proceed only if calculation was successful
        st.header(f"Calculation Results (in {display_currency})")

        # Update top display line to include rebate % if applicable
        display_destination = f" | Dest: `{selected_destination}`" if selected_shipment_type == "Air" else ""
        display_pallet = f" | Pallet: `{selected_pallet_type}` (x{num_pallets})" if selected_pallet_type not in [None, "None"] else ""
        display_rebate = f" | Rebate: `{rebate_percentage:.1f}%`" if rebate_percentage > 0 else "" # Add rebate display
        st.write(
            f"**Product:** `{selected_product_str}` | **Qty:** `{quantity_input}` | **Ship Via:** `{selected_shipment_type}`"
            f"{display_destination}{display_pallet}{display_rebate} | **Fixed Costs:** `{fixed_cost_selection}`" # Added rebate here
        )

        colw1, colw2, colw3 = st.columns(3)
        colw1.metric("Net Wt/Box", f"{net_weight_kg:.3f} KG")
        colw2.metric("Gross Wt/Box", f"{calculated_gross_weight_kg_per_box:.3f} KG")
        colw3.metric("Final Shipping Wt (Batch)", f"{final_shipping_gross_weight_kg:.3f} KG", f"{num_pallets} Pallets")

        st.markdown("---")
        # Renamed Delivered Cost tab
        tab_cogs, tab_logistics, tab_total, tab_summary, tab_profit = st.tabs([
            "💰 COGS",
            f"🚚 Logistics ({selected_shipment_type})",
            "📦 Delivered Cost (+Rebate)", # Renamed
            "📊 Batch Summary Detail",
            "📈 Profit Analysis"
        ])

        # --- COGS Tab ---
        # ... (remains the same) ...
        with tab_cogs:
            st.subheader(f"Cost of Goods Sold {fixed_cost_label_suffix}")
            col1a, col1b = st.columns(2); col1a.metric("Total COGS", format_cost_by_mode(total_cogs_usd, currency_display_mode)); col1b.metric("COGS / Box", format_cost_by_mode(cogs_per_box_usd, currency_display_mode)); col1b.metric("COGS / KG (Net)", format_cost_by_mode(cogs_per_kg_usd, currency_display_mode))
            st.subheader("COGS Components (Total)"); col2a, col2b, col2c = st.columns(3); col2a.metric("Raw Product", format_cost_by_mode(total_raw_cost_usd, currency_display_mode)); col2b.metric("Variable (incl. Pallets)", format_cost_by_mode(total_variable_costs_incl_pallets_usd, currency_display_mode), help=f"Only included if checkbox is checked. Pallet cost always included."); col2c.metric(f"Fixed Costs {fixed_cost_label_suffix}", format_cost_by_mode(total_allocated_fixed_cost_usd, currency_display_mode), help=f"Includes Calc. Interest: {format_cost_by_mode(interest_cost_usd, currency_display_mode)}" if interest_cost_usd > 0 else None)
            st.subheader("COGS Breakdown (Per Box)"); st.write(f"- Raw Product: {format_cost_by_mode(raw_cost_per_box_usd, currency_display_mode)}"); st.write(f"- Variable (incl. Pallets): {format_cost_by_mode(variable_costs_incl_pallets_per_box_usd, currency_display_mode)}"); st.caption(f"  (Components: {format_cost_by_mode(total_per_unit_variable_comp_cost_usd, currency_display_mode)} + Pallets: {format_cost_by_mode(pallet_cost_per_box_usd, currency_display_mode)})"); st.write(f"- Fixed Costs {fixed_cost_label_suffix}: {format_cost_by_mode(fixed_cost_per_unit_usd, currency_display_mode)}"); st.write(f"**= Total COGS Per Box:** {format_cost_by_mode(cogs_per_box_usd, currency_display_mode)}")

        # --- Logistics Tab ---
        # ... (remains the same) ...
        with tab_logistics:
            st.subheader(f"{selected_shipment_type} Logistics Cost ({logistics_cost_source})"); col3a, col3b = st.columns(2); col3a.metric("Total Logistics Cost", format_cost_by_mode(total_logistics_cost_usd, currency_display_mode)); col3b.metric("Logistics / Box", format_cost_by_mode(logistics_per_box_usd, currency_display_mode));
            if final_shipping_gross_weight_kg > 0: col3b.metric("Logistics / KG (Gross Ship Wt)", format_cost_by_mode(logistics_per_kg_gross_usd, currency_display_mode))
            st.subheader("Logistics Components (Total)");
            if selected_shipment_type == "Air": col4a, col4b = st.columns(2); freight_cost = final_shipping_gross_weight_kg * logistics_rate_per_kg; col4a.metric("Freight Cost", format_cost_by_mode(freight_cost, currency_display_mode), f"{final_shipping_gross_weight_kg:.2f}kg @ {format_cost_by_mode(logistics_rate_per_kg, currency_display_mode)}/kg"); col4b.metric("Airway Bill Cost", format_cost_by_mode(awb_cost, currency_display_mode))
            elif selected_shipment_type in ["Container", "Truck"]:
                st.metric(f"{selected_shipment_type} Fixed Price", format_cost(fixed_logistics_price))
            else: st.write("N/A")

        # --- Total Delivered Cost Tab (Modified for Rebate) ---
        with tab_total:
            st.subheader("Total Delivered Cost (COGS + Logistics + Unexpected + Rebate)") # Updated title
            # Show costs *before* rebate first
            st.metric("Total Delivered Cost (Before Rebate)", format_cost_by_mode(total_delivered_cost_usd, currency_display_mode))
            st.write(f"*Calculation: {format_cost_by_mode(total_cogs_usd, currency_display_mode)} (COGS) + {format_cost_by_mode(total_logistics_cost_usd, currency_display_mode)} (Logistics) + {format_cost_by_mode(total_unexpected_cost_usd, currency_display_mode)} (Unexpected)*")
            st.markdown("---")
            # Show rebate calculation
            st.subheader(f"Retailer Rebate/Fee Adjustment ({rebate_percentage:.1f}%)")
            col_rebate1, col_rebate2 = st.columns(2)
            col_rebate1.metric("Rebate/Fee Amount (Total Batch)", format_cost_by_mode(rebate_amount_usd, currency_display_mode))
            # Show final costs *after* rebate
            col_rebate2.metric("Final Total Cost (Incl. Rebate)", format_cost_by_mode(final_total_cost_usd, currency_display_mode))
            col_rebate2.metric("Final Cost / Box (Incl. Rebate)", format_cost_by_mode(final_cost_per_box_usd, currency_display_mode))
            if total_net_weight_kg > 0:
                 final_cost_per_kg_net_usd = final_total_cost_usd / total_net_weight_kg
                 col_rebate2.metric("Final Cost / KG Net (Incl. Rebate)", format_cost_by_mode(final_cost_per_kg_net_usd, currency_display_mode))
            st.write(f"*Final Calculation: {format_cost_by_mode(total_delivered_cost_usd, currency_display_mode)} (Delivered Cost) + {format_cost_by_mode(rebate_amount_usd, currency_display_mode)} (Rebate)*")


        # --- Batch Summary Detail Tab (Modified for Rebate) ---
        with tab_summary:
            st.subheader("Total Batch Cost Summary")
            summary_data_dict = st.session_state.get('summary_data', {})
            if summary_data_dict:
                # Pie Chart: Shows breakdown BEFORE rebate for clarity on operational costs
                plot_data = {
                    'Raw Product': total_raw_cost_usd,
                    'Variable (incl. Pallets)': total_variable_costs_incl_pallets_usd,
                    'Fixed Costs': total_allocated_fixed_cost_usd,
                    'Freight/Fee': freight_or_fixed_logistics_cost,
                    'Unexpected': total_unexpected_cost_usd
                }
                plot_data_filtered = {k: v for k, v in plot_data.items() if v > 0}
                if plot_data_filtered:
                    plot_df = pd.DataFrame(list(plot_data_filtered.items()), columns=pd.Index(["Cost Category", "Total Cost (USD)"]))
                    _ledger_palette = ["#e11d74", "#6b3fd4", "#b87d00", "#0b8f7e", "#15121f", "#a8a0b0"]
                    fig = px.pie(plot_df, names='Cost Category', values='Total Cost (USD)',
                                 title='Cost breakdown by component (before rebate)',
                                 hole=.58,
                                 color_discrete_sequence=_ledger_palette)
                    fig.update_traces(textposition='outside', textinfo='percent+label',
                                      marker=dict(line=dict(color="#ffffff", width=2)))
                    fig.update_layout(
                        paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
                        font=dict(family="Space Grotesk, sans-serif", color="#15121f", size=13),
                        title=dict(font=dict(family="Space Grotesk, sans-serif", size=16, color="#15121f"), x=0, xanchor="left"),
                        legend=dict(font=dict(family="JetBrains Mono, monospace", size=11)),
                        margin=dict(l=10, r=10, t=60, b=10)
                    )
                    st.plotly_chart(fig, use_container_width=True)
                else: st.write("No primary cost components > 0 for chart.")

                # Summary Table: Shows the FULL breakdown including rebate and final total
                st.subheader("Full Cost Breakdown Table")
                summary_df = pd.DataFrame(list(summary_data_dict.items()), columns=pd.Index(["Cost Category", "Total Cost (USD/EUR)"]))
                st.dataframe(summary_df, use_container_width=True)
            else: st.write("Summary data not available (Run calculation first).")


        # --- Profit Analysis Tab ---
        with tab_profit:
            st.subheader("Profit Margin Analysis")
            
            # Get stored profit analysis data
            profit_data = st.session_state.get('profit_analysis', {})
            if profit_data:
                final_cost_per_box = profit_data.get('final_cost_per_box_usd', 0)
                cogs_per_box = profit_data.get('cogs_per_box_usd', 0)
                delivered_cost_per_box = profit_data.get('delivered_cost_per_box_usd', 0)
                
                # Display current costs
                col_cost1, col_cost2, col_cost3 = st.columns(3)
                col_cost1.metric("COGS per Box", format_cost_by_mode(cogs_per_box, currency_display_mode))
                col_cost2.metric("Delivered Cost per Box", format_cost_by_mode(delivered_cost_per_box, currency_display_mode))
                col_cost3.metric("Final Cost per Box (incl. Rebate)", format_cost_by_mode(final_cost_per_box, currency_display_mode))
                
                st.markdown("---")
                
                # Interactive Profit Calculator
                st.subheader("Profit Calculator")
                sales_price_input = st.number_input(
                    "Enter Sales Price per Box (USD):",
                    min_value=0.0,
                    value=final_cost_per_box * 1.2,  # Default 20% markup
                    step=0.01,
                    format="%.2f"
                )
                
                if sales_price_input > 0:
                    # Calculate profit margins
                    profit_analysis = calculate_profit_margins(final_cost_per_box, sales_price_input)
                    
                    col_profit1, col_profit2, col_profit3 = st.columns(3)
                    col_profit1.metric("Profit per Box", format_cost_by_mode(profit_analysis['profit_per_box'], currency_display_mode))
                    col_profit2.metric("Profit Margin", f"{profit_analysis['profit_margin_percent']:.1f}%")
                    col_profit3.metric("ROI", f"{profit_analysis['roi_percent']:.1f}%")
                    
                    # Total profit for batch
                    total_profit = profit_analysis['profit_per_box'] * quantity_input
                    st.metric("Total Profit for Batch", format_cost_by_mode(total_profit, currency_display_mode))
                    
                    # Cost Sensitivity Analysis
                    st.markdown("---")
                    st.subheader("Cost Sensitivity Analysis")
                    
                    # Create sensitivity table
                    markup_percentages = [5, 10, 15, 20, 25, 30, 40, 50]
                    sensitivity_data = []
                    
                    for markup in markup_percentages:
                        sales_price = final_cost_per_box * (1 + markup / 100)
                        analysis = calculate_profit_margins(final_cost_per_box, sales_price)
                        sensitivity_data.append({
                            "Markup %": markup,
                            "Sales Price": format_cost_by_mode(sales_price, currency_display_mode),
                            "Profit per Box": format_cost_by_mode(analysis['profit_per_box'], currency_display_mode),
                            "Profit Margin %": f"{analysis['profit_margin_percent']:.1f}%",
                            "ROI %": f"{analysis['roi_percent']:.1f}%"
                        })
                    
                    sensitivity_df = pd.DataFrame(sensitivity_data)
                    st.dataframe(sensitivity_df, use_container_width=True)

                    # Add a graph for Cost Sensitivity Analysis
                    fig = go.Figure()
                    _line_palette = ["#e11d74", "#6b3fd4", "#0b8f7e"]
                    def _coerce_num(s):
                        return float(str(s).replace('$','').replace(',','').replace('€','').replace('£','').replace('₺','').replace('%','').replace('C$','').replace('A$','').replace('S$','').replace('CHF ','').replace('¥','').replace('د.إ ','').strip())
                    fig.add_trace(go.Scatter(x=sensitivity_df['Markup %'], y=sensitivity_df['Profit per Box'].apply(_coerce_num), mode='lines+markers', name='Profit per Box', line=dict(color=_line_palette[0], width=2.2), marker=dict(size=7)))
                    fig.add_trace(go.Scatter(x=sensitivity_df['Markup %'], y=sensitivity_df['Profit Margin %'].apply(_coerce_num), mode='lines+markers', name='Profit Margin %', line=dict(color=_line_palette[1], width=2.2), marker=dict(size=7)))
                    fig.add_trace(go.Scatter(x=sensitivity_df['Markup %'], y=sensitivity_df['ROI %'].apply(_coerce_num), mode='lines+markers', name='ROI %', line=dict(color=_line_palette[2], width=2.2), marker=dict(size=7)))
                    fig.update_layout(
                        title=dict(text='Cost sensitivity', font=dict(family="Space Grotesk, sans-serif", size=16, color="#15121f"), x=0, xanchor="left"),
                        xaxis=dict(title='Markup %', gridcolor="rgba(20,16,30,0.08)", zerolinecolor="rgba(20,16,30,0.08)", linecolor="#d9d5e0"),
                        yaxis=dict(title='Value', gridcolor="rgba(20,16,30,0.08)", zerolinecolor="rgba(20,16,30,0.08)", linecolor="#d9d5e0"),
                        paper_bgcolor="#ffffff", plot_bgcolor="#ffffff",
                        font=dict(family="Space Grotesk, sans-serif", color="#15121f", size=13),
                        legend=dict(font=dict(size=11)), margin=dict(l=20, r=20, t=60, b=40)
                    )
                    st.plotly_chart(fig, use_container_width=True)
                    
            else:
                st.info("Run a calculation first to see profit analysis.")

        # --- Optional Variable Breakdown ---
        with st.expander("Show Detailed Variable Cost & Weight Breakdown (Per Box) - Components Only"):
             # Ensure product_variable_costs_detailed is available and has the cost column if needed
            if not product_variable_costs_detailed.empty and 'LineItemCost_USD' in product_variable_costs_detailed.columns:
                 # Select columns carefully in case Cost was excluded
                 cols_to_show = ['ComponentName', 'QuantityPerProduct', 'WeightKG', 'LineItemWeightKG']
                 if include_variable_costs:
                     cols_to_show.extend(['CostPerUnit_USD', 'LineItemCost_USD'])

                 breakdown_display = product_variable_costs_detailed[cols_to_show]
                 # Rename columns for display
                 rename_map = {
                     'CostPerUnit_USD': 'Cost/Unit ($)', 'LineItemCost_USD': 'TotalCost ($)',
                     'WeightKG': 'Wt/Unit (KG)', 'LineItemWeightKG': 'TotalWt (KG)'
                 }
                 breakdown_display = breakdown_display.rename(columns=rename_map)

                 # Define formatters, applying only if column exists
                 formatters = {
                    'QuantityPerProduct': '{:,.2f}',
                    'Wt/Unit (KG)': '{:,.4f}', 'TotalWt (KG)': '{:,.4f}'
                 }
                 if include_variable_costs:
                     formatters['Cost/Unit ($)'] = '${:,.6f}'
                     formatters['TotalCost ($)'] = '${:,.6f}'
                 st.dataframe(breakdown_display.set_index('ComponentName').style.format(formatters))
                 st.write(f"**Total Component Packaging Weight (per Box):** {total_packaging_weight_kg:.4f} KG")
                 if not include_variable_costs:
                     st.caption("*Variable Component Costs were excluded from COGS calculation based on checkbox selection.*")
            else: st.write("(No variable components found in recipe or breakdown unavailable)")

# --- Export Section ---
st.markdown("---")
st.subheader("📊 Export Results")

if st.session_state.get('calculation_done', False):
    # Get data for export
    summary_data_dict = st.session_state.get('summary_data', {})
    profit_data = st.session_state.get('profit_analysis', {})
    calculation_details = {
        'product': st.session_state.get('last_calc_product', 'N/A'),
        'quantity': st.session_state.get('last_calc_quantity', 'N/A'),
        'shipment_type': st.session_state.get('selected_shipment_type', 'N/A'),
        'fixed_cost_mode': fixed_cost_mode if 'fixed_cost_mode' in locals() else 'N/A'
    }
    
    # Create sensitivity data for export
    sensitivity_data = []
    if profit_data and 'final_cost_per_box_usd' in profit_data:
        final_cost_per_box = profit_data['final_cost_per_box_usd']
        markup_percentages = [5, 10, 15, 20, 25, 30, 40, 50]
        for markup in markup_percentages:
            sales_price = final_cost_per_box * (1 + markup / 100)
            analysis = calculate_profit_margins(final_cost_per_box, sales_price)
            sensitivity_data.append({
                "Markup %": markup,
                "Sales Price (USD)": f"${sales_price:,.2f}",
                "Profit per Box (USD)": f"${analysis['profit_per_box']:,.2f}",
                "Profit Margin %": f"{analysis['profit_margin_percent']:.1f}%",
                "ROI %": f"{analysis['roi_percent']:.1f}%"
            })
    
    col_export1, col_export2 = st.columns(2)
    
    with col_export1:
        st.write("**Export as CSV:**")
        csv_data = create_csv_export(summary_data_dict, profit_data, calculation_details)
        csv_filename = f"cost_calculation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        csv_link = get_download_link(csv_data, csv_filename, "csv")
        st.markdown(csv_link, unsafe_allow_html=True)
    
    with col_export2:
        st.write("**Export as Excel:**")
        excel_data = create_excel_export(summary_data_dict, profit_data, calculation_details, sensitivity_data)
        excel_filename = f"cost_calculation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        excel_link = get_download_link(excel_data, excel_filename, "excel")
        st.markdown(excel_link, unsafe_allow_html=True)
    
    st.caption("💡 Export includes cost breakdown, profit analysis, and sensitivity analysis (Excel only)")
else:
    st.info("Run a calculation first to enable export options.")

# --- UI Improvements and Advanced Features ---
st.markdown("---")
st.subheader("🎨 Advanced Features")

# Cost History Tracking
col_theme2 = st.columns(1)[0]
with col_theme2:
    st.write("**Cost History:**")
    if st.button("💾 Save Current Calculation", help="Save this calculation to history"):
        if st.session_state.get('calculation_done', False):
            # Save calculation to session state history
            if 'cost_history' not in st.session_state:
                st.session_state['cost_history'] = []
            
            history_entry = {
                'timestamp': datetime.now().isoformat(),
                'product': st.session_state.get('last_calc_product', 'N/A'),
                'quantity': st.session_state.get('last_calc_quantity', 'N/A'),
                'final_cost_per_box': st.session_state.get('final_cost_per_box_usd', 0),
                'shipment_type': st.session_state.get('selected_shipment_type', 'N/A'),
                'fixed_cost_mode': fixed_cost_mode if 'fixed_cost_mode' in locals() else 'N/A'
            }
            st.session_state['cost_history'].append(history_entry)
            st.success("Calculation saved to history!")

# Display Cost History
if 'cost_history' in st.session_state and st.session_state['cost_history']:
    st.subheader("📈 Cost History")
    history_df = pd.DataFrame(st.session_state['cost_history'])
    history_df['timestamp'] = pd.to_datetime(history_df['timestamp']).dt.strftime('%Y-%m-%d %H:%M')
    history_df['final_cost_per_box'] = history_df['final_cost_per_box'].apply(lambda x: format_cost_by_mode(x, currency_display_mode))
    
    # Rename columns for display
    history_df = history_df.rename(columns={
        'timestamp': 'Date/Time',
        'product': 'Product',
        'quantity': 'Quantity',
        'final_cost_per_box': 'Cost per Box',
        'shipment_type': 'Shipment',
        'fixed_cost_mode': 'Fixed Cost Mode'
    })
    
    st.dataframe(history_df, use_container_width=True)
    
    # Clear history button
    if st.button("🗑️ Clear History"):
        st.session_state['cost_history'] = []
        st.rerun()

# --- Input prompts if calculation isn't ready ---
if not st.session_state.get('calculation_done', False):
    # Check specific conditions to give more targeted advice
    if not selected_product: st.warning("Select a product to begin.")
    elif quantity_input < 1: st.warning("Enter a valid Quantity (>= 1).")
    elif selected_shipment_type == "Select...": st.warning("Select a shipment type.")
    elif selected_shipment_type == "Air" and (selected_destination is None or selected_destination == "Select..."): st.warning("Select a destination for Air shipment.")
    elif selected_pallet_type is None: st.warning("Select a pallet type (use 'None' if no pallets).")
    else: st.info("Adjust inputs in the sidebar and click 'Calculate Costs'.") # Default message


# --- Display raw data ---
# ... (remains the same) ...
st.markdown("---")
with st.expander("Show Raw Data Loaded from Files"):
    # Display DataFrames safely checking if they exist
    if pallets_df is not None: st.write(f"**Pallet Specifications (`{PALLETS_CSV}`):**"); st.dataframe(pallets_df.style.format({'CostUSD': '${:,.2f}', 'WeightKG': '{:,.2f} KG'})); st.write("---")
    if air_rates_df is not None: st.write(f"**Air Freight Rates (`{AIR_RATES_CSV}` - USD):**"); st.dataframe(air_rates_df.style.format({'MinWeightKG': '{:,.1f} KG','PricePerKG_USD': '${:,.3f}','AirwayBill_USD': '${:,.2f}'})); st.write("---")
    if product_weights_df is not None: st.write(f"**Product Net Weights (`{WEIGHTS_CSV}`):**"); st.dataframe(product_weights_df.style.format({'NetWeightKG': '{:.3f} KG'})); st.write("---")
    if product_packing_df is not None: st.write(f"**Product Packing (`{PACKING_CSV}`):**"); st.dataframe(product_packing_df.style.format({'BoxesPerPallet': '{:,.0f}'})); st.write("---") # Added Packing DF
    if components_df is not None: st.write(f"**Processed Components (`{COMPONENTS_CSV}` - USD Cost, KG Weight):**"); st.dataframe(components_df[['ComponentName', 'CostPerUnit_USD', 'WeightKG']].style.format({'CostPerUnit_USD': '${:,.6f}', 'WeightKG': '{:,.4f} KG'})); st.write("---")
    if product_recipe_df is not None: st.write(f"**Product Recipes (`{RECIPE_CSV}`):**"); st.dataframe(product_recipe_df); st.write("---")
    if fixed_df is not None: st.write(f"**Processed Fixed Costs (`{FIXED_CSV}` - USD Monthly):**"); st.dataframe(fixed_df[['CostItem', 'MonthlyCost_USD', 'Category']].style.format({'MonthlyCost_USD': '${:,.2f}'}))
