import io

import pandas as pd
from openpyxl import load_workbook

from cogs.exporters import season_matrix_to_excel


def _long():
    return pd.DataFrame(
        [
            {"Produce": "Cherry", "Pack type": "Cherry 10 x 500g", "Boxes/pallet": 75,
             "Destination": "BAH-TK", "2 pallets": 46.81, "4 pallets": 40.19,
             "6 pallets": 38.42, "10 pallets": 37.01},
            {"Produce": "Cherry", "Pack type": "Cherry 10 x 500g", "Boxes/pallet": 75,
             "Destination": "HKG-TK", "2 pallets": 46.15, "4 pallets": 39.54,
             "6 pallets": 37.77, "10 pallets": 36.35},
        ]
    )


def test_season_matrix_to_excel_roundtrips():
    data = season_matrix_to_excel(_long(), price_basis="Cost", built_at="2026-06-08 12:00")
    assert isinstance(data, bytes) and data[:2] == b"PK"  # .xlsx is a zip archive

    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Season pricing", "Run info"]

    ws = wb["Season pricing"]
    assert [c.value for c in ws[1]] == [
        "Produce", "Pack type", "Boxes/pallet", "Destination",
        "2 pallets", "4 pallets", "6 pallets", "10 pallets",
    ]
    assert ws.cell(row=2, column=5).value == 46.81           # first "2 pallets" price
    assert "$" in ws.cell(row=2, column=5).number_format     # currency-formatted
    assert ws.freeze_panes == "A2"                           # header frozen
    assert ws.max_row == 3                                   # header + 2 data rows


def test_season_matrix_to_excel_records_inputs():
    inputs = pd.DataFrame({"Produce": ["Cherry"], "Raw TRY/kg": [115.5], "Raw USD/kg": [3.5]})
    data = season_matrix_to_excel(_long(), price_basis="Sell price", built_at="x", inputs_df=inputs)
    flat = [c.value for row in load_workbook(io.BytesIO(data))["Run info"].iter_rows() for c in row]
    assert "Sell price" in flat
    assert "Cherry" in flat and 115.5 in flat


def test_season_matrix_to_excel_empty_is_safe():
    empty = pd.DataFrame(columns=["Produce", "Pack type", "Destination"])
    wb = load_workbook(io.BytesIO(season_matrix_to_excel(empty)))
    assert "Season pricing" in wb.sheetnames
    assert wb["Season pricing"].max_row == 1  # header only
